# actualizar.py — Dimensión 1: Ser joven
# Script para procesar las proyecciones de población de Bogotá (DANE, CNPV 2018)
# y generar los JSONs que usa la página web.
#
# Uso:
#   1. Descargar el anexo de proyecciones de Bogotá del DANE:
#      https://www.dane.gov.co/index.php/estadisticas-por-tema/demografia-y-poblacion/proyecciones-de-poblacion/proyecciones-de-poblacion-bogota
#
#   2. Guardar el archivo en dim1/fuentes/
#
#   3. Correr este script:
#      python dim1/actualizar.py
#
#   4. Los archivos JSON en dim1/data/ se actualizan automáticamente.

import os
import json
import openpyxl

# ============================================================
# Configuración
# ============================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FUENTES_DIR = os.path.join(SCRIPT_DIR, 'fuentes')
DATA_DIR = os.path.join(SCRIPT_DIR, 'data')

# Rango de edad para "joven" según Ley 1622 (Estatuto de Ciudadanía Juvenil)
EDAD_MIN = 14
EDAD_MAX = 28

# Años a incluir en el JSON (no necesitamos todo 2018-2035 para el tablero)
ANIO_MIN = 2018
ANIO_MAX = 2035

# Nombre de la hoja principal en el Excel del DANE
HOJA_PRINCIPAL = 'Localidades'

# Fila donde empiezan los headers (puede cambiar si el DANE modifica el formato)
FILA_HEADERS = 12


# ============================================================
# Funciones
# ============================================================

def encontrar_archivo_proyecciones(carpeta):
    """Busca el archivo de proyecciones más reciente en la carpeta."""
    archivos = [f for f in os.listdir(carpeta)
                if f.endswith('.xlsx') and not f.startswith('~$')
                and 'proyecciones' in f.lower()]
    if not archivos:
        print(f'ERROR: No se encontró ningún archivo de proyecciones en {carpeta}')
        return None
    archivos.sort()
    return os.path.join(carpeta, archivos[-1])


def leer_headers(ws, fila):
    """Lee los nombres de columna de la fila de headers."""
    headers = []
    for cell in ws[fila]:
        headers.append(cell.value)
    return headers


def extraer_datos(ws, headers):
    """Extrae todos los datos de la hoja en una lista de diccionarios."""
    # Índices de columnas clave
    idx_loc = headers.index('COD_LOC')
    idx_nom = headers.index('NOM_LOC')
    idx_area = headers.index('AREA') if 'AREA' in headers else headers.index('ÁREA')
    # El año puede venir como AÑO o A?O por encoding
    idx_anio = None
    for i, h in enumerate(headers):
        if h and ('AÑO' in str(h).upper() or 'A\ufffdO' in str(h) or 'ANO' in str(h).upper()
                  or str(h).upper().startswith('A') and str(h).upper().endswith('O') and len(str(h)) <= 4):
            idx_anio = i
            break
    if idx_anio is None:
        idx_anio = 3  # posición por defecto

    # Mapear columnas de edad por sexo
    # Formato: Hombres_14, Mujeres_14, Total_14
    col_hombres = {}  # edad -> índice de columna
    col_mujeres = {}
    col_total = {}

    for i, h in enumerate(headers):
        if not h:
            continue
        h = str(h)
        for prefix, destino in [('Hombres_', col_hombres), ('Mujeres_', col_mujeres), ('Total_', col_total)]:
            if h.startswith(prefix):
                edad_str = h[len(prefix):]
                if edad_str.isdigit():
                    destino[int(edad_str)] = i

    # Índices de totales generales
    idx_total_h = None
    idx_total_m = None
    idx_total = None
    for i, h in enumerate(headers):
        if h == 'TOTAL HOMBRES':
            idx_total_h = i
        elif h == 'TOTAL MUJERES':
            idx_total_m = i
        elif h == 'TOTAL':
            idx_total = i

    registros = []
    for row in ws.iter_rows(min_row=FILA_HEADERS + 1, values_only=True):
        if not row[idx_loc]:
            continue

        anio = int(row[idx_anio]) if row[idx_anio] else None
        if not anio or anio < ANIO_MIN or anio > ANIO_MAX:
            continue

        # Población joven por edad y sexo
        jovenes_h = {}
        jovenes_m = {}
        jovenes_t = {}
        for edad in range(EDAD_MIN, EDAD_MAX + 1):
            if edad in col_hombres:
                jovenes_h[edad] = int(row[col_hombres[edad]] or 0)
            if edad in col_mujeres:
                jovenes_m[edad] = int(row[col_mujeres[edad]] or 0)
            if edad in col_total:
                jovenes_t[edad] = int(row[col_total[edad]] or 0)

        total_jovenes_h = sum(jovenes_h.values())
        total_jovenes_m = sum(jovenes_m.values())
        total_jovenes = sum(jovenes_t.values())
        total_poblacion = int(row[idx_total] or 0) if idx_total else 0
        total_pob_h = int(row[idx_total_h] or 0) if idx_total_h else 0
        total_pob_m = int(row[idx_total_m] or 0) if idx_total_m else 0

        registros.append({
            'cod_loc': str(row[idx_loc]).zfill(2),
            'localidad': row[idx_nom],
            'area': row[idx_area],
            'anio': anio,
            'jovenes_hombres': total_jovenes_h,
            'jovenes_mujeres': total_jovenes_m,
            'jovenes_total': total_jovenes,
            'poblacion_total': total_poblacion,
            'poblacion_hombres': total_pob_h,
            'poblacion_mujeres': total_pob_m,
            'por_edad_hombres': jovenes_h,
            'por_edad_mujeres': jovenes_m,
        })

    return registros


def generar_jsons(registros):
    """Genera los archivos JSON para el tablero web."""
    os.makedirs(DATA_DIR, exist_ok=True)

    # 1. Resumen por año para Bogotá (suma de todas las localidades, area=Total)
    resumen = {}
    for r in registros:
        if r['area'] != 'Total':
            continue
        anio = r['anio']
        if anio not in resumen:
            resumen[anio] = {
                'anio': anio,
                'jovenes_hombres': 0, 'jovenes_mujeres': 0, 'jovenes_total': 0,
                'poblacion_total': 0,
                'por_edad_hombres': {}, 'por_edad_mujeres': {},
                'zona_cabecera': 0, 'zona_rural': 0,
            }
        resumen[anio]['jovenes_hombres'] += r['jovenes_hombres']
        resumen[anio]['jovenes_mujeres'] += r['jovenes_mujeres']
        resumen[anio]['jovenes_total'] += r['jovenes_total']
        resumen[anio]['poblacion_total'] += r['poblacion_total']
        # Acumular por edad
        for edad in range(EDAD_MIN, EDAD_MAX + 1):
            e_str = str(edad)
            resumen[anio]['por_edad_hombres'][e_str] = (
                resumen[anio]['por_edad_hombres'].get(e_str, 0) + r['por_edad_hombres'].get(edad, 0)
            )
            resumen[anio]['por_edad_mujeres'][e_str] = (
                resumen[anio]['por_edad_mujeres'].get(e_str, 0) + r['por_edad_mujeres'].get(edad, 0)
            )

    # Agregar zona (cabecera vs rural) desde los registros por área
    for r in registros:
        anio = r['anio']
        if anio not in resumen:
            continue
        if r['area'] == 'Cabecera Municipal':
            resumen[anio]['zona_cabecera'] += r['jovenes_total']
        elif r['area'] == 'Centros Poblados y Rural Disperso':
            resumen[anio]['zona_rural'] += r['jovenes_total']

    resumen_list = sorted(resumen.values(), key=lambda x: x['anio'])

    ruta_resumen = os.path.join(DATA_DIR, 'resumen_bogota.json')
    with open(ruta_resumen, 'w', encoding='utf-8') as f:
        json.dump(resumen_list, f, ensure_ascii=False, indent=2)
    print(f'  resumen_bogota.json: {len(resumen_list)} años')

    # 2. Datos por localidad (area=Total, solo totales de jóvenes)
    localidades = []
    for r in registros:
        if r['area'] != 'Total':
            continue
        localidades.append({
            'cod_loc': r['cod_loc'],
            'localidad': r['localidad'],
            'anio': r['anio'],
            'jovenes_hombres': r['jovenes_hombres'],
            'jovenes_mujeres': r['jovenes_mujeres'],
            'jovenes_total': r['jovenes_total'],
            'poblacion_total': r['poblacion_total'],
        })

    ruta_loc = os.path.join(DATA_DIR, 'localidades.json')
    with open(ruta_loc, 'w', encoding='utf-8') as f:
        json.dump(localidades, f, ensure_ascii=False, indent=2)
    locs_unicas = len(set(r['localidad'] for r in localidades))
    print(f'  localidades.json: {len(localidades)} registros ({locs_unicas} localidades)')

    return ruta_resumen, ruta_loc


def main():
    print('=' * 60)
    print('Actualización de datos — Dimensión 1: Ser joven')
    print('=' * 60)

    print(f'\nBuscando archivo de proyecciones en: {FUENTES_DIR}')
    archivo = encontrar_archivo_proyecciones(FUENTES_DIR)
    if not archivo:
        return
    print(f'  Archivo: {os.path.basename(archivo)}')

    print('\nLeyendo datos del DANE...')
    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    ws = wb[HOJA_PRINCIPAL]

    headers = leer_headers(ws, FILA_HEADERS)
    registros = extraer_datos(ws, headers)
    wb.close()
    print(f'  {len(registros)} registros extraídos')

    print('\nGenerando archivos JSON...')
    ruta_resumen, ruta_loc = generar_jsons(registros)

    # Verificación rápida
    with open(ruta_resumen, 'r', encoding='utf-8') as f:
        resumen = json.load(f)
    ultimo = [r for r in resumen if r['anio'] == 2024]
    if ultimo:
        d = ultimo[0]
        print(f'\n{"=" * 60}')
        print(f'VERIFICACIÓN — Año 2024:')
        print(f'{"=" * 60}')
        print(f'  Jóvenes (14-28): {d["jovenes_total"]:,.0f}'.replace(',', '.'))
        pct = d['jovenes_total'] / d['poblacion_total'] * 100 if d['poblacion_total'] else 0
        print(f'  % de la población: {pct:.2f}%'.replace('.', ','))
        print(f'  Zona cabecera: {d["zona_cabecera"]:,.0f}'.replace(',', '.'))
        print(f'  Zona rural: {d["zona_rural"]:,.0f}'.replace(',', '.'))
        print(f'  Hombres: {d["jovenes_hombres"]:,.0f}'.replace(',', '.'))
        print(f'  Mujeres: {d["jovenes_mujeres"]:,.0f}'.replace(',', '.'))

    print(f'\nArchivos actualizados:')
    print(f'  - {ruta_resumen}')
    print(f'  - {ruta_loc}')
    print(f'\nEl tablero web se actualizará automáticamente al hacer push.')


if __name__ == '__main__':
    main()
