# actualizar_datos.py
# Script para actualizar los datos del tablero de Inclusión Productiva
# cuando se descarga un nuevo anexo del DANE (GEIH Mercado laboral juvenil)
#
# Uso:
#   1. Descargar el anexo más reciente del DANE:
#      https://www.dane.gov.co/index.php/estadisticas-por-tema/mercado-laboral/mercado-laboral-de-la-juventud
#      (archivo tipo "anex-GEIHMLJ-xxx-xxx20XX.xlsx")
#
#   2. Guardar el archivo en:
#      240708_Fuentes/Actualizacion 2026/Mercado laboral/
#
#   3. Correr este script desde la carpeta tablero-cij:
#      python actualizar_datos.py
#
#   4. Verificar en http://localhost:8080 (python -m http.server 8080)
#
#   5. Si todo se ve bien, hacer push a GitHub:
#      git add data/
#      git commit -m "Actualizar datos mercado laboral [trimestre]"
#      git push

import os
import re
import json
import openpyxl

# ============================================================
# Configuración - ajustar si cambian las rutas
# ============================================================

# Carpeta raíz del proyecto
TABLERO_DIR = os.path.dirname(os.path.abspath(__file__))

# Carpeta de fuentes del Power BI
FUENTES_DIR = os.path.join(
    os.path.dirname(TABLERO_DIR),
    'Tableros power bi', 'Tableros',
    'Visualización Centro de Información de Juventud',
    '240708_Fuentes', 'Actualizacion 2026', 'Mercado laboral'
)

# Archivo de salida del Power BI (se actualiza también)
PBI_FILE_NAME = 'Mercado laboral jóvenes.xlsx'

# Carpeta de datos del tablero web
DATA_DIR = os.path.join(TABLERO_DIR, 'data')

# Ciudades a extraer para comparación
CIUDADES_EXTRAER = {
    'Bogotá D.C.': ('23 ciudades trim móvil', 29),
    'Medellín A.M.': ('23 ciudades trim móvil', 46),
    'Cali A.M.': ('23 ciudades trim móvil', 63),
    'Barranquilla A.M.': ('23 ciudades trim móvil', 80),
    'Bucaramanga A.M.': ('23 ciudades trim móvil', 97),
    'Total nacional': (' Tnal trimestre móvil', 12),
    'Total 13 ciudades': ('13 ciudades trimestre móvil', 12),
}


# ============================================================
# Funciones
# ============================================================

def encontrar_anexo_mas_reciente(carpeta):
    """Busca el archivo anex-GEIHMLJ más reciente en la carpeta."""
    archivos = [f for f in os.listdir(carpeta) if f.startswith('anex-GEIHMLJ') and f.endswith('.xlsx')]
    if not archivos:
        print(f'ERROR: No se encontró ningún archivo anex-GEIHMLJ*.xlsx en {carpeta}')
        return None
    # Ordenar por nombre (el más reciente tiene la fecha más alta)
    archivos.sort()
    return os.path.join(carpeta, archivos[-1])


def normalizar_trimestre(trim):
    """Normaliza espacios y capitalización de un nombre de trimestre."""
    t = re.sub(r'\s+', ' ', trim).strip()
    parts = t.split(' ')
    normalized = []
    for p in parts:
        if p and p[0].isalpha():
            normalized.append(p[0].upper() + p[1:])
        else:
            normalized.append(p)
    return ' '.join(normalized)


def extraer_ciudad(ws, start_row, anio_min=2014):
    """Extrae datos de una sección de ciudad en la hoja del DANE."""
    all_rows = []
    for i, row in enumerate(ws.iter_rows(min_row=start_row+1, max_row=start_row+14, values_only=True)):
        all_rows.append(list(row))

    if len(all_rows) < 14:
        return []

    year_row = all_rows[1]
    trim_row = all_rows[2]

    current_year = None
    col_map = {}
    for col_idx, val in enumerate(year_row):
        if val is not None and isinstance(val, (int, float)):
            current_year = int(val)
        if current_year and col_idx < len(trim_row) and trim_row[col_idx]:
            col_map[col_idx] = (current_year, trim_row[col_idx])

    results = []
    for col_idx, (year, trim) in col_map.items():
        if year < anio_min:
            continue

        def get_val(row_idx, col):
            if col < len(all_rows[row_idx]):
                v = all_rows[row_idx][col]
                if v is not None:
                    try:
                        return round(float(v), 1)
                    except (ValueError, TypeError):
                        pass
            return None

        d = {
            'anio': year,
            'trimestre': normalizar_trimestre(trim),
            'pct_pet': get_val(3, col_idx),
            'tgp': get_val(4, col_idx),
            'to': get_val(5, col_idx),
            'td': get_val(6, col_idx),
            'pct_fuera_ft': get_val(7, col_idx),
            'pet': get_val(8, col_idx),
            'pet_15_28': get_val(9, col_idx),
            'fuerza_trabajo': get_val(10, col_idx),
            'ocupados': get_val(11, col_idx),
            'desocupados': get_val(12, col_idx),
            'fuera_ft': get_val(13, col_idx),
        }
        if d['tgp'] is not None:
            results.append(d)

    return results


def actualizar_pbi_excel(pbi_path, bogota_data):
    """Actualiza el archivo Excel del Power BI con los datos nuevos de Bogotá."""
    # Crear backup
    backup_path = pbi_path.replace('.xlsx', '_backup.xlsx')
    import shutil
    shutil.copy2(pbi_path, backup_path)
    print(f'  Backup creado: {os.path.basename(backup_path)}')

    wb = openpyxl.load_workbook(pbi_path)
    ws = wb['Mercado laboral']

    # Leer encabezados existentes (fila 1)
    headers = [ws.cell(row=1, column=c).value for c in range(1, 14)]

    # Encontrar última fila con datos
    last_row = 1
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        if row[0].value is not None:
            last_row = row[0].row

    # Mapeo de columnas DANE → Excel
    col_map = {
        'anio': 0, 'trimestre': 1, 'pct_pet': 2, 'tgp': 3, 'to': 4,
        'td': 5, 'pct_fuera_ft': 6, 'pet': 7, 'pet_15_28': 8,
        'fuerza_trabajo': 9, 'ocupados': 10, 'desocupados': 11, 'fuera_ft': 12
    }

    # Leer datos existentes para detectar duplicados
    existing = set()
    for row in ws.iter_rows(min_row=2, max_row=last_row, max_col=2, values_only=True):
        if row[0] and row[1]:
            existing.add((row[0], normalizar_trimestre(str(row[1]))))

    # Agregar datos nuevos
    nuevos = 0
    for d in bogota_data:
        key = (d['anio'], d['trimestre'])
        if key not in existing:
            last_row += 1
            vals = [d['anio'], d['trimestre'], d['pct_pet'], d['tgp'], d['to'],
                    d['td'], d['pct_fuera_ft'],
                    round(d['pet']) if d['pet'] else None,
                    round(d['pet_15_28']) if d['pet_15_28'] else None,
                    round(d['fuerza_trabajo']) if d['fuerza_trabajo'] else None,
                    round(d['ocupados']) if d['ocupados'] else None,
                    round(d['desocupados']) if d['desocupados'] else None,
                    round(d['fuera_ft']) if d['fuera_ft'] else None]
            for col_idx, val in enumerate(vals, 1):
                ws.cell(row=last_row, column=col_idx, value=val)
            nuevos += 1

    wb.save(pbi_path)
    wb.close()
    print(f'  Excel PBI actualizado: {nuevos} trimestres nuevos agregados')


def main():
    print('=' * 60)
    print('Actualización de datos - Mercado laboral juvenil')
    print('=' * 60)

    # 1. Encontrar el anexo DANE más reciente
    print(f'\nBuscando anexo DANE en: {FUENTES_DIR}')
    anexo_path = encontrar_anexo_mas_reciente(FUENTES_DIR)
    if not anexo_path:
        return
    print(f'  Archivo encontrado: {os.path.basename(anexo_path)}')

    # 2. Abrir el archivo
    print('\nLeyendo datos del DANE...')
    wb = openpyxl.load_workbook(anexo_path, read_only=True, data_only=True)

    # 3. Extraer datos de todas las ciudades
    all_cities = {}
    for city_name, (sheet_name, start_row) in CIUDADES_EXTRAER.items():
        ws = wb[sheet_name]
        anio_min = 2014 if city_name == 'Bogotá D.C.' else 2018
        data = extraer_ciudad(ws, start_row, anio_min=anio_min)
        all_cities[city_name] = data
        ultimo = data[-1] if data else None
        print(f'  {city_name}: {len(data)} registros'
              f' (último: {ultimo["trimestre"] if ultimo else "?"})')
    wb.close()

    # 4. Guardar JSON para el tablero web
    print('\nGenerando archivos JSON...')
    os.makedirs(DATA_DIR, exist_ok=True)

    # mercado_laboral.json (solo Bogotá, todos los trimestres desde 2014)
    bogota_data = all_cities['Bogotá D.C.']
    bogota_path = os.path.join(DATA_DIR, 'mercado_laboral.json')
    with open(bogota_path, 'w', encoding='utf-8') as f:
        json.dump(bogota_data, f, ensure_ascii=False, indent=2)
    print(f'  mercado_laboral.json: {len(bogota_data)} registros')

    # mercado_laboral_ciudades.json (ciudades comparación, desde 2018)
    ciudades_data = {}
    for city, data in all_cities.items():
        ciudades_data[city] = [d for d in data if d['anio'] >= 2018]
    ciudades_path = os.path.join(DATA_DIR, 'mercado_laboral_ciudades.json')
    with open(ciudades_path, 'w', encoding='utf-8') as f:
        json.dump(ciudades_data, f, ensure_ascii=False, indent=2)
    print(f'  mercado_laboral_ciudades.json: {len(ciudades_data)} ciudades')

    # 5. Actualizar Excel del Power BI
    pbi_path = os.path.join(FUENTES_DIR, PBI_FILE_NAME)
    if os.path.exists(pbi_path):
        print(f'\nActualizando Excel Power BI: {PBI_FILE_NAME}')
        actualizar_pbi_excel(pbi_path, bogota_data)
    else:
        print(f'\nNota: No se encontró {PBI_FILE_NAME} en la carpeta de fuentes.')
        print('  El Excel del Power BI no se actualizó.')

    # 6. Resumen
    fijos = ['Ene - Mar', 'Abr - Jun', 'Jul - Sep', 'Oct - Dic']
    print('\n' + '=' * 60)
    print('RESUMEN')
    print('=' * 60)
    print(f'Fuente: {os.path.basename(anexo_path)}')
    print(f'Trimestres fijos disponibles por año (Bogotá):')
    for year in sorted(set(d['anio'] for d in bogota_data)):
        if year < 2018:
            continue
        trims = [d['trimestre'] for d in bogota_data
                 if d['anio'] == year and any(d['trimestre'].startswith(f) for f in fijos)]
        print(f'  {year}: {", ".join(trims)}')

    print(f'\nArchivos actualizados:')
    print(f'  - {bogota_path}')
    print(f'  - {ciudades_path}')
    if os.path.exists(pbi_path):
        print(f'  - {pbi_path}')

    print(f'\nSiguientes pasos:')
    print(f'  1. Verificar en http://localhost:8080')
    print(f'     (correr: python -m http.server 8080)')
    print(f'  2. git add data/')
    print(f'     git commit -m "Actualizar datos mercado laboral"')
    print(f'     git push')


if __name__ == '__main__':
    main()
